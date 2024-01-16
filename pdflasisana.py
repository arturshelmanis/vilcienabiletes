import PyPDF2
import pathlib
from openpyxl import Workbook, load_workbook

data = []
adrese = pathlib.Path("Mail_Attachments")
visi_faili = list(adrese.glob("*.pdf"))
viss_kopa = 0

for f in range(len(visi_faili)):

    pdf_file = PyPDF2.PdfReader(open(visi_faili[f], "rb"))
    page = pdf_file.pages[0]
    text = page.extract_text()

    start_position = text.find("Kopā apmaksai") + len("Kopā apmaksai") + 1 #viss teksts + vēl viena atstarpe
    end_postition = text.find("Tai skaitā PVN")
    kopeja_summa = text[start_position : end_postition].rstrip()
    kopeja_summa = kopeja_summa.replace("€", "")
    kopeja_summa = kopeja_summa.strip()
    
    if text.find("(līdz 23:59)") != -1:
        start_position = text.find("(līdz 23:59)") + len("(līdz 23:59)") + 1
    else:
        start_position = text.find("Biļetes derīguma datums") + len("Biļetes derīguma datums") + 1
    
    end_postition = start_position + 10
    datums = text[start_position : end_postition].rstrip()

    #print(f"Datums: {datums}, summma: {kopeja_summa}")
    data.append([datums, kopeja_summa])
    viss_kopa += 1
    #print(text)
    
print(viss_kopa)

wb = load_workbook('source_col.xlsx')
ws = wb.active 



months_list = [ ["01.2023", "Janvāris"], ["02.2023", "Februāris"], ["03.2023", "Marts"], ["04.2023", "Aprīlis"], ["05.2023", "Maijs"], ["06.2023", "Jūnijs"],
           ["07.2023", "Jūlijs"], ["08.2023", "Augusts"], ["09.2023", "Septembris"], ["10.2023", "Oktobris"], ["11.2023", "Novembris"], ["12.2023", "Decembris"], ["01.2024", "Janvāris, 2024"] ]
count = 1
collumn_char = 65
for month in months_list:
    if count > 1:
        collumn_char = collumn_char + 1
    count = 1
    for ticket in data:
        datums = ticket[0]
        kopeja_summa = ticket[1]

        if (datums[3:10] == month[0]):
            if count == 1:
                ws[str(chr(collumn_char)) + str(count)].value = month[1]
            count += 1
            ws[str(chr(collumn_char)) + str(count)].value = float(ticket[1])
wb.save("vilcienaizmaksas.xlsx")
wb.close()


