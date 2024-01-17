import imaplib
import email
import os
from datetime import datetime
import pandas as pd
import PyPDF2
import pathlib
import time
from openpyxl import Workbook, load_workbook
from user_login import *

folder_name = "Vilciena_Biletes"
if not os.path.exists(folder_name):
    os.makedirs(folder_name, mode=0o777)
folder_path = os.path.abspath(folder_name)

imap_url = 'mail.inbox.lv'
my_mail = imaplib.IMAP4_SSL(imap_url)
my_mail.login(my_email, password_key)

my_mail.select("INBOX")
key = "FROM"
value = "info@mobilly.lv"

from_which = str(input("Ievadiet no kura datuma vēlaties iegūt datus: (DD.MM.YYYY) "))
from_which =  datetime.strptime(from_which, "%d.%m.%Y")
from_which =  datetime.strftime(from_which, "%d-%b-%Y")

to_which = str(input("Ievadiet līdz kuram datumam vēlaties iegūt datus: (DD.MM.YYYY) "))
to_which =  datetime.strptime(to_which, "%d.%m.%Y")
to_which =  datetime.strftime(to_which, "%d-%b-%Y")

print(f"Tiek uzsākta informācijas iegūšana par laika periodu no {from_which} līdz {to_which}!")

aaa, data = my_mail.search(None, key, value, f'(SINCE "{from_which}")', f'(BEFORE "{to_which}")')
list_of_tickets = data[0].split()

count = 1
for i in list_of_tickets:
    typ, data = my_mail.fetch(i, '(RFC822)')
    mail=data[0][1]
    msg = email.message_from_string(mail.decode('utf-8'))

    for part in msg.walk():
        if part.get_content_maintype() == 'multipart':
                continue
        if part.get('Content-Disposition') is None:
                continue
        
        if count < 10:
            count = "0" + str(count)

        filename = "Bilete_" + str(count)
        if filename:
            filepath = os.path.join(folder_path,filename)
            with open(f"{filepath}.pdf", 'wb') as f:
                f.write(part.get_payload(decode=True))
        
        count = int(count)
        count += 1

print("Iegūtās biļetes par laika periodu tika saglabātas mapē Vilciena_Biletes")
my_mail.logout()

#Sākas PDF failu lasīšana

from_which =  datetime.strptime(from_which, "%d-%b-%Y")
to_which =  datetime.strptime(to_which, "%d-%b-%Y")

months_list =pd.date_range(from_which, to_which, freq='MS').strftime("%m.%Y|%B").tolist()

adrese = pathlib.Path("Vilciena_Biletes")
visi_faili = list(adrese.glob("*.pdf"))
viss_kopa = 0
data = []
for f in range(len(visi_faili)):

    pdf_file = PyPDF2.PdfReader(open(visi_faili[f], "rb"))
    page = pdf_file.pages[0]
    text = page.extract_text()

    start_position = text.find("Kopā apmaksai") + len("Kopā apmaksai") + 1 #viss teksts + vēl viena atstarpe
    end_postition = text.find("Tai skaitā PVN")
    kopeja_summa = text[start_position : end_postition].rstrip()
    kopeja_summa = kopeja_summa.replace("€", "")
    kopeja_summa = kopeja_summa.strip()
    
    if text.find("(līdz 23:59)") != -1:
        start_position = text.find("(līdz 23:59)") + len("(līdz 23:59)") + 1
    else:
        start_position = text.find("Biļetes derīguma datums") + len("Biļetes derīguma datums") + 1
    
    end_postition = start_position + 10
    datums = text[start_position : end_postition].rstrip()

    #print(f"Datums: {datums}, summma: {kopeja_summa}")
    data.append([datums, kopeja_summa])
    viss_kopa += 1
    #print(text)
    
print(viss_kopa)

wb = Workbook()
ws = wb.active 

count = 1
collumn_char = 65
for month in months_list:
    if count > 1:
        collumn_char = collumn_char + 1
    count = 1
    for ticket in data:
        datums = ticket[0]
        kopeja_summa = ticket[1]

        if (datums[3:10] == month[0:7]):
            if count == 1:
                ws[str(chr(collumn_char)) + str(count)].value = month[8:] + ", " + month[3:7]
            count += 1
            ws[str(chr(collumn_char)) + str(count)].value = float(ticket[1])

wb.save("vilcienaizmaksas.xlsx")
wb.close()